from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox

# Função para ler os dados da planilha
def ler_dados(Inventario_Onze):
    try:
        workbook = load_workbook(filename=Inventario_Onze)
        sheet = workbook.active

        dados = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            item = {
                'Nome': row[0],
                'OC/OG': row[1],
                'OLD USER': row[2],
                'FABRICANTE': row[3],
                'MODELO': row[4],
                'S/N': row[5],
                'ATIVO': row[6],
                'CARREGADOR': row[7],
                'VENDOR': row[8],
                'NF': row[9],
                'DATA': row[10],
                'Valor Unitário': row[11],
                'Valor Residual': row[12],
                'Status de Venda': row[13]
            }
            dados.append(item)

        return dados
    except FileNotFoundError:
        print("Arquivo não encontrado.")

# Função para atualizar a planilha com os dados modificados
def atualizar_planilha(Inventario_Onze, dados):
    try:
        workbook = load_workbook(filename=Inventario_Onze)
        sheet = workbook.active

        # Limpar a planilha
        sheet.delete_rows(2, sheet.max_row)

        # Preencher com os novos dados
        for idx, item in enumerate(dados):
            row = idx + 2
            sheet.cell(row=row, column=1).value = item['Nome']
            sheet.cell(row=row, column=2).value = item['OC/OG']
            sheet.cell(row=row, column=3).value = item['OLD USER']
            sheet.cell(row=row, column=4).value = item['FABRICANTE']
            sheet.cell(row=row, column=5).value = item['MODELO']
            sheet.cell(row=row, column=6).value = item['S/N']
            sheet.cell(row=row, column=7).value = item['ATIVO']
            sheet.cell(row=row, column=8).value = item['CARREGADOR']
            sheet.cell(row=row, column=9).value = item['VENDOR']
            sheet.cell(row=row, column=10).value = item['NF']
            sheet.cell(row=row, column=11).value = item['DATA']
            sheet.cell(row=row, column=12).value = item['Valor Unitário']
            sheet.cell(row=row, column=13).value = item['Valor Residual']
            sheet.cell(row=row, column=14).value = item['Status de Venda']

        workbook.save(Inventario_Onze)
        messagebox.showinfo("Sucesso", "Alterações salvas com sucesso!")
    except:
        messagebox.showerror("Erro", "Erro ao salvar as alterações.")

# Função para adicionar um novo item à planilha
def adicionar_item():
    item = {
        'Nome': entry_nome.get(),
        'OC/OG': entry_ocog.get(),
        'OLD USER': entry_olduser.get(),
        'FABRICANTE': entry_fabricante.get(),
        'MODELO': entry_modelo.get(),
        'S/N': entry_sn.get(),
        'ATIVO': entry_ativo.get(),
        'CARREGADOR': entry_carregador.get(),
        'VENDOR': entry_vendor.get(),
        'NF': entry_nf.get(),
        'DATA': entry_data.get(),
        'Valor Unitário': entry_valorunitario.get(),
        'Valor Residual': entry_valorresidual.get(),
        'Status de Venda': entry_statusvenda.get()
    }
    dados.append(item)
    atualizar_tabela()
    messagebox.showinfo("Sucesso", "Item adicionado com sucesso.")

# Função para remover um item da planilha
def remover_item():
    if len(lista_itens.curselection()) > 0:
        index = lista_itens.curselection()[0]
        del dados[index]
        atualizar_tabela()
        messagebox.showinfo("Sucesso", "Item removido com sucesso.")
    else:
        messagebox.showwarning("Atenção", "Nenhum item selecionado.")

# Função para atualizar a tabela com os dados
def atualizar_tabela():
    lista_itens.delete(0, tk.END)
    for item in dados:
        lista_itens.insert(tk.END, item['Nome'])

# Função para exibir os
        from tkinter import Tk, Frame, Scrollbar, Listbox, Label, Entry, Button

# Função para exibir os detalhes do item selecionado
def exibir_detalhes(event):
    if len(lista_itens.curselection()) > 0:
        index = lista_itens.curselection()[0]
        item = dados[index]
        entry_nome.delete(0, 'end')
        entry_nome.insert('end', item['Nome'])
        entry_ocog.delete(0, 'end')
        entry_ocog.insert('end', item['OC/OG'])
        # Continuar com os outros campos

# Função para atualizar os detalhes do item selecionado
def atualizar_detalhes():
    if len(lista_itens.curselection()) > 0:
        index = lista_itens.curselection()[0]
        item = dados[index]
        item['Nome'] = entry_nome.get()
        item['OC/OG'] = entry_ocog.get()
        # Continuar com os outros campos
        atualizar_planilha(Inventario_Onze, dados)
        messagebox.showinfo("Sucesso", "Detalhes atualizados com sucesso.")

# Criação da janela principal
root = Tk()
root.title("Inventário TI")

# Frame para a lista de itens
frame_lista = Frame(root)
frame_lista.pack(side="left", fill="y")

# Scrollbar para a lista de itens
scrollbar = Scrollbar(frame_lista, orient="vertical")
scrollbar.pack(side="right", fill="y")

# Lista de itens
lista_itens = Listbox(frame_lista, yscrollcommand=scrollbar.set)
lista_itens.pack(side="left", fill="both")

# Configuração da scrollbar
scrollbar.config(command=lista_itens.yview)

# Frame para os detalhes do item selecionado
frame_detalhes = Frame(root)
frame_detalhes.pack(side="right", fill="both")

# Labels e Entries para os detalhes do item selecionado
label_nome = Label(frame_detalhes, text="Nome:")
label_nome.pack()
entry_nome = Entry(frame_detalhes)
entry_nome.pack()

label_ocog = Label(frame_detalhes, text="OC/OG:")
label_ocog.pack()
entry_ocog = Entry(frame_detalhes)
entry_ocog.pack()

# Continuar com os outros campos

# Botão para atualizar os detalhes do item
botao_atualizar = Button(frame_detalhes, text="Atualizar", command=atualizar_detalhes)
botao_atualizar.pack()

# Leitura dos dados da planilha
dados = ler_dados(Inventario_Onze)

# Atualização da tabela
atualizar_tabela()

# Configuração do evento de seleção de item na lista
lista_itens.bind('<<ListboxSelect>>', exibir_detalhes)

# Execução da janela
root.mainloop()